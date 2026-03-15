"""Multi-format document ingestors: PDF (PyMuPDF+pytesseract), Email (MBOX/EML), CSV/Excel."""
from __future__ import annotations

import csv
import email
import io
import mailbox
from dataclasses import dataclass, field
from pathlib import Path

from rich.console import Console

console = Console()


@dataclass
class IngestedDocument:
    """Normalised document ready for EvidenceClient.ingest_document()."""
    content: str
    source_type: str  # "PDF", "EMAIL", "CSV", "EXCEL", "WEB"
    source_url: str   # file path or original URL
    matter_id: str
    metadata: dict = field(default_factory=dict)
    page_number: int | None = None
    message_id: str | None = None


# ─────────────────────────────────────────────
# PDF Ingestor
# ─────────────────────────────────────────────

class PDFIngestor:
    """Extract text from PDF using PyMuPDF (fitz). Falls back to pytesseract OCR for image pages."""

    def ingest(self, file_path: str | Path, matter_id: str) -> list[IngestedDocument]:
        """Extract text from all pages. Returns one IngestedDocument per page.
        
        Uses fitz (PyMuPDF) for text extraction.
        Falls back to pytesseract OCR if page text is empty (scanned page).
        """
        try:
            import fitz  # PyMuPDF
        except ImportError as exc:
            raise ImportError("PyMuPDF (fitz) is required for PDF ingestion. Install: pip install pymupdf") from exc

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"PDF file not found: {file_path}")

        documents = []
        try:
            doc = fitz.open(str(file_path))
            console.print(f"[cyan]PDF: {file_path.name} — {doc.page_count} pages[/cyan]")

            for page_num in range(doc.page_count):
                page = doc[page_num]
                text = page.get_text("text").strip()

                # Fallback to OCR for image-only pages
                if not text:
                    text = self._ocr_page(page)

                if not text:
                    continue

                documents.append(IngestedDocument(
                    content=text,
                    source_type="PDF",
                    source_url=str(file_path),
                    matter_id=matter_id,
                    page_number=page_num + 1,
                    metadata={
                        "filename": file_path.name,
                        "page": page_num + 1,
                        "total_pages": doc.page_count,
                        "ocr_used": not bool(page.get_text("text").strip()),
                    },
                ))

            doc.close()
        except Exception as exc:
            console.print(f"[red]PDF ingestion failed for {file_path}: {exc}[/red]")
            raise

        console.print(f"[green]PDF: extracted {len(documents)} pages from {file_path.name}[/green]")
        return documents

    def _ocr_page(self, page) -> str:
        """OCR a fitz page using pytesseract."""
        try:
            import pytesseract
            from PIL import Image
            # Render page to image at 300 DPI
            mat = page.get_pixmap(dpi=300)
            img_data = mat.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            text = pytesseract.image_to_string(img)
            return text.strip()
        except ImportError:
            console.print("[yellow]pytesseract/Pillow not available — skipping OCR for image page[/yellow]")
            return ""
        except Exception as exc:
            console.print(f"[yellow]OCR failed: {exc}[/yellow]")
            return ""


# ─────────────────────────────────────────────
# Email Ingestor
# ─────────────────────────────────────────────

class EmailIngestor:
    """Extract emails from MBOX files or individual EML files."""

    def ingest_mbox(self, file_path: str | Path, matter_id: str) -> list[IngestedDocument]:
        """Extract all emails from an MBOX file."""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"MBOX file not found: {file_path}")

        documents = []
        try:
            mbox = mailbox.mbox(str(file_path))
            count = 0
            for msg in mbox:
                doc = self._parse_email_message(msg, str(file_path), matter_id)
                if doc:
                    documents.append(doc)
                    count += 1
            console.print(f"[green]MBOX: extracted {count} emails from {file_path.name}[/green]")
        except Exception as exc:
            console.print(f"[red]MBOX ingestion failed for {file_path}: {exc}[/red]")
            raise

        return documents

    def ingest_eml(self, file_path: str | Path, matter_id: str) -> IngestedDocument | None:
        """Extract a single EML file."""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"EML file not found: {file_path}")

        try:
            with open(file_path, "rb") as f:
                msg = email.message_from_bytes(f.read())
            doc = self._parse_email_message(msg, str(file_path), matter_id)
            return doc
        except Exception as exc:
            console.print(f"[red]EML ingestion failed for {file_path}: {exc}[/red]")
            raise

    def _parse_email_message(self, msg, source_path: str, matter_id: str) -> IngestedDocument | None:
        """Parse an email.message.Message into IngestedDocument."""
        try:
            subject = str(msg.get("Subject", "") or "")
            sender = str(msg.get("From", "") or "")
            recipient = str(msg.get("To", "") or "")
            date_str = str(msg.get("Date", "") or "")
            message_id = str(msg.get("Message-ID", "") or "").strip()

            # Extract body text
            body = self._extract_email_body(msg)
            if not body and not subject:
                return None

            content = f"Subject: {subject}\nFrom: {sender}\nTo: {recipient}\nDate: {date_str}\n\n{body}"

            return IngestedDocument(
                content=content,
                source_type="EMAIL",
                source_url=source_path,
                matter_id=matter_id,
                message_id=message_id or None,
                metadata={
                    "subject": subject,
                    "from": sender,
                    "to": recipient,
                    "date": date_str,
                    "message_id": message_id,
                },
            )
        except Exception as exc:
            console.print(f"[yellow]Failed to parse email message: {exc}[/yellow]")
            return None

    def _extract_email_body(self, msg) -> str:
        """Extract plain text body from email, preferring text/plain over text/html."""
        body_parts = []
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                disposition = str(part.get("Content-Disposition", ""))
                if "attachment" in disposition:
                    continue
                if content_type == "text/plain":
                    try:
                        charset = part.get_content_charset() or "utf-8"
                        body_parts.append(part.get_payload(decode=True).decode(charset, errors="replace"))
                    except Exception:
                        pass
        else:
            try:
                charset = msg.get_content_charset() or "utf-8"
                payload = msg.get_payload(decode=True)
                if payload:
                    body_parts.append(payload.decode(charset, errors="replace"))
            except Exception:
                pass

        return "\n".join(body_parts).strip()


# ─────────────────────────────────────────────
# CSV / Excel Ingestor
# ─────────────────────────────────────────────

class SpreadsheetIngestor:
    """Extract structured data from CSV and Excel files."""

    ROWS_PER_CHUNK = 100  # Rows per IngestedDocument chunk

    def ingest_csv(self, file_path: str | Path, matter_id: str) -> list[IngestedDocument]:
        """Read CSV and produce one IngestedDocument per ROWS_PER_CHUNK rows."""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"CSV file not found: {file_path}")

        documents = []
        try:
            with open(file_path, newline="", encoding="utf-8-sig", errors="replace") as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames or []
                rows_buffer: list[dict] = []
                chunk_index = 0

                for row in reader:
                    rows_buffer.append(dict(row))
                    if len(rows_buffer) >= self.ROWS_PER_CHUNK:
                        doc = self._rows_to_document(
                            rows_buffer, headers, file_path, matter_id, chunk_index, "CSV"
                        )
                        documents.append(doc)
                        rows_buffer = []
                        chunk_index += 1

                if rows_buffer:
                    doc = self._rows_to_document(
                        rows_buffer, headers, file_path, matter_id, chunk_index, "CSV"
                    )
                    documents.append(doc)

        except Exception as exc:
            console.print(f"[red]CSV ingestion failed for {file_path}: {exc}[/red]")
            raise

        console.print(f"[green]CSV: {len(documents)} chunks from {file_path.name}[/green]")
        return documents

    def ingest_excel(self, file_path: str | Path, matter_id: str) -> list[IngestedDocument]:
        """Read Excel (.xlsx/.xls) and produce one IngestedDocument per sheet per ROWS_PER_CHUNK rows."""
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl is required for Excel ingestion. Install: pip install openpyxl") from exc

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        documents = []
        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                headers = [str(h) if h is not None else "" for h in rows[0]]
                data_rows = rows[1:]
                chunk_index = 0

                for i in range(0, len(data_rows), self.ROWS_PER_CHUNK):
                    chunk = data_rows[i : i + self.ROWS_PER_CHUNK]
                    row_dicts = [
                        {headers[j]: str(cell) if cell is not None else "" for j, cell in enumerate(row)}
                        for row in chunk
                    ]
                    doc = self._rows_to_document(
                        row_dicts, headers, file_path, matter_id, chunk_index, "EXCEL",
                        extra_meta={"sheet": sheet_name},
                    )
                    documents.append(doc)
                    chunk_index += 1

            wb.close()
        except Exception as exc:
            console.print(f"[red]Excel ingestion failed for {file_path}: {exc}[/red]")
            raise

        console.print(f"[green]Excel: {len(documents)} chunks from {file_path.name}[/green]")
        return documents

    def _rows_to_document(
        self,
        rows: list[dict],
        headers: list,
        file_path: Path,
        matter_id: str,
        chunk_index: int,
        source_type: str,
        extra_meta: dict | None = None,
    ) -> IngestedDocument:
        """Convert a list of row dicts to an IngestedDocument."""
        # Represent rows as readable text: "header: value | header: value"
        lines = []
        for row in rows:
            parts = [f"{k}: {v}" for k, v in row.items() if v not in (None, "", "None")]
            if parts:
                lines.append(" | ".join(parts))

        content = "\n".join(lines)
        meta = {
            "filename": file_path.name,
            "chunk_index": chunk_index,
            "row_count": len(rows),
            "headers": list(headers),
        }
        if extra_meta:
            meta.update(extra_meta)

        return IngestedDocument(
            content=content,
            source_type=source_type,
            source_url=str(file_path),
            matter_id=matter_id,
            metadata=meta,
        )


# ─────────────────────────────────────────────
# Unified ingestor factory
# ─────────────────────────────────────────────

def ingest_file(file_path: str | Path, matter_id: str) -> list[IngestedDocument]:
    """Auto-detect file type and ingest. Returns list of IngestedDocument."""
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".pdf":
        return PDFIngestor().ingest(file_path, matter_id)
    elif suffix in (".mbox",):
        return EmailIngestor().ingest_mbox(file_path, matter_id)
    elif suffix in (".eml", ".msg"):
        doc = EmailIngestor().ingest_eml(file_path, matter_id)
        return [doc] if doc else []
    elif suffix == ".csv":
        return SpreadsheetIngestor().ingest_csv(file_path, matter_id)
    elif suffix in (".xlsx", ".xls"):
        return SpreadsheetIngestor().ingest_excel(file_path, matter_id)
    else:
        raise ValueError(f"Unsupported file type: {suffix}. Supported: .pdf, .mbox, .eml, .csv, .xlsx, .xls")
